from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd 
from datetime import date, datetime
import openpyxl
from io import BytesIO
import os
from pytz import timezone


app = Flask(__name__)

@app.route("/")
def salary_app():
    return render_template("index.html", status = "", salary="", year = datetime.now(timezone('EST')).year, date = datetime.now(timezone('EST')).strftime("%m/%d/%Y"))

@app.route('/process_amount', methods=['POST'])
def process_amount():
    # Access the submitted data from the HTML form
    
    amount = request.form['amount']
    bincome = 'bincome' in request.form
    bexpense = 'bexpense' in request.form
    bmakeup = 'bmakeup' in request.form
    best = 'best' in request.form
    bbar = 'bbar' in request.form
    bcheer = 'bcheer' in request.form
    date_ = request.form['date']
    
    status = enter_amount(amount, bincome, bmakeup, best, bbar, bcheer, date_)
    
    # Redirect or render another template as needed
    return render_template("index.html", status = status, salary="", year = datetime.now(timezone('EST')).year, date = datetime.now(timezone('EST')).strftime("%m/%d/%Y"))

@app.route('/process_salary', methods=['POST'])
def process_salary():
    # Access the submitted data from the HTML form
    bnet = 'bnet' in request.form
    bgross = 'bgross' in request.form

    bmakeup = 'bmakeup2' in request.form
    best = 'best2' in request.form
    bbar = 'bbar2' in request.form
    bcheer = 'bcheer2' in request.form
    
    year = request.form['year']
    
    salary = calc_salary(bnet, bgross, bmakeup, best, bbar, bcheer, year)
    
    # Redirect or render another template as needed
    return render_template("index.html", status = "", salary=salary, year = datetime.now(timezone('EST')).year, date = datetime.now(timezone('EST')).strftime("%m/%d/%Y"))

@app.route('/display_data')
def display_data():
    # Read data from the Excel file
    workbook = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), 'salary_data.xlsx'))
    try:
        sheet = workbook['Sheet1']
    except:
        sheet = workbook['Sheet']

    # Get header and data
    header = [cell.value for cell in sheet[1]]
    header = header[0:]
    data = [[cell.value for cell in row][0:] for row in sheet.iter_rows(min_row=2)]

    return render_template('table.html', header=header, data=data)

@app.route('/remove_rows', methods=['POST'])
def remove_rows():
    # Get the list of row indices to remove from the request
    rows_to_remove = request.json.get('rowsToRemove', [])

    # Read data from the Excel file
    workbook = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), 'salary_data.xlsx'))
    try:
        sheet = workbook['Sheet1']
    except:
        sheet = workbook['Sheet']

    # Remove selected rows
    for i in reversed(rows_to_remove):
        sheet.delete_rows(i + 2)  # Adding 2 to account for 1-based index and header row

    # Save the updated workbook
    updated_data = BytesIO()
    workbook.save(updated_data)
    updated_data.seek(0)

    return send_file(updated_data, download_name='salary_data.xlsx', as_attachment=True,  mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download_excel', methods=['POST'])
def download_excel():
    # Get the data from the request
    data = request.json.get('data', [])
    # Create a new workbook and write the data to it
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write header
    header = data[0]
    sheet.append(header)

    # Write data
    for row_data in data[1:]:
        sheet.append(row_data)

    # Save the workbook to a BytesIO object
    updated_data = BytesIO()
    workbook.save(updated_data)
    updated_data.seek(0)

    # Send the file in the response
    return send_file(updated_data, download_name='salary_data.xlsx', as_attachment=True,  mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/overwrite_excel', methods=['POST'])
def overwrite_excel():
    
    # Get the JSON data from the request
    data = request.json.get('data', [])

    # Call the function to save the Excel file
    save_excel(data)

    return jsonify(success=True)


def enter_amount(amount, bincome, bmakeup, best, bbar, bcheer, date_):
        
    if not amount:
        return ""

    amount = float(amount)
    salary_df = pd.read_excel(os.path.join(os.path.dirname(__file__), 'salary_data.xlsx'))

    try:
        salary_df.drop(columns={'Unnamed: 0'}, inplace=True)
    except:
        pass

    if not bincome:
        amount = amount*-1

    salary_dict = {'entries': [0], 'makeup': [0], 'bartending': [0], 'esthetician': [0], 'cheer': [0], 'date': [0], 'year': []}
    if bmakeup:
        salary_dict['makeup'] = [amount]
    elif bbar:
        salary_dict['bartending'] = [amount]
    elif best:
        salary_dict['esthetician'] = [amount]
    elif bcheer:
        salary_dict['cheer'] = [amount]

    salary_dict['year'] = datetime.now(timezone('EST')).year
    salary_dict['date'] = str(date_)
    salary_dict['entries'] = len(salary_df)+1

    new_df = pd.DataFrame(salary_dict)
    salary_df = pd.concat([salary_df, new_df])
    salary_df.reset_index(drop=True, inplace=True)

    try:
        salary_df.drop(columns={'Unnamed: 0'}, inplace=True)
    except:
        pass

    salary_df.to_excel(os.path.join(os.path.dirname(__file__), 'salary_data.xlsx'), index=False)

    return "Entry Added"

def calc_salary(bnet, bgross, bmakeup, best, bbar, bcheer, year):
        
    if not bnet and not bgross:
        return "Select Net or Gross"
    
    year = int(year)
    salary_df = pd.read_excel(os.path.join(os.path.dirname(__file__), 'salary_data.xlsx'))

    try:
        salary_df.drop(columns={'Unnamed: 0'}, inplace=True)
    except:
        pass

    if bnet:
        salary_df = salary_df[salary_df['year'] == year]
        if (bmakeup and best and bbar and bcheer) or (not bmakeup and not best and not bbar and not bcheer):
            salary = salary_df['makeup'].sum() + salary_df['bartending'].sum() + salary_df['esthetician'].sum() + salary_df['cheer'].sum()

        elif (bmakeup and best and bcheer and not bbar):
            salary = salary_df['makeup'].sum() + salary_df['esthetician'].sum() + salary_df['cheer'].sum()
        elif (bmakeup and bbar and bcheer and not best):
            salary = salary_df['makeup'].sum() + salary_df['bartending'].sum() + salary_df['cheer'].sum()
        elif (best and bbar and bcheer and not bmakeup):
            salary = salary_df['esthetician'].sum() + salary_df['bartending'].sum() + salary_df['cheer'].sum()
        elif (best and bbar and bmakeup and not bcheer):
            salary = salary_df['esthetician'].sum() + salary_df['bartending'].sum() + salary_df['makeup'].sum()

        elif (bmakeup and best and not bcheer and not bbar):
            salary = salary_df['makeup'].sum() + salary_df['esthetician'].sum() 
        elif (bmakeup and bbar and not bcheer and not best):
            salary = salary_df['makeup'].sum() + salary_df['bartending'].sum()
        elif (not best and not bbar and bcheer and bmakeup):
            salary = salary_df['makeup'].sum() + salary_df['cheer'].sum()

        elif (best and not bbar and bcheer and not bmakeup):
            salary = salary_df['esthetician'].sum() + salary_df['cheer'].sum()
        elif (best and bbar and not bcheer and not bmakeup):
            salary = salary_df['esthetician'].sum() + salary_df['bartending'].sum()

        elif (not best and bbar and bcheer and not bmakeup):
            salary = salary_df['bartending'].sum() + salary_df['cheer'].sum()

        elif (bmakeup and not best and not bbar and not bcheer):
            salary = salary_df['makeup'].sum()
        elif (best and not bbar and not bmakeup and not bcheer):
            salary = salary_df['esthetician'].sum()
        elif (bbar and not best and not bmakeup and not bcheer):
            salary = salary_df['bartending'].sum()
        elif (bcheer and not bbar and not best and not bmakeup):
            salary = salary_df['cheer'].sum()
    elif bgross:
        salary_df = salary_df[salary_df['year'] == year]
        salary_df = salary_df[salary_df['makeup'] >= 0]
        salary_df = salary_df[salary_df['bartending'] >= 0]
        salary_df = salary_df[salary_df['esthetician'] >= 0]

        if (bmakeup and best and bbar and bcheer) or (not bmakeup and not best and not bbar and not bcheer):
            salary = salary_df['makeup'].sum() + salary_df['bartending'].sum() + salary_df['esthetician'].sum() + salary_df['cheer'].sum()

        elif (bmakeup and best and bcheer and not bbar):
            salary = salary_df['makeup'].sum() + salary_df['esthetician'].sum() + salary_df['cheer'].sum()
        elif (bmakeup and bbar and bcheer and not best):
            salary = salary_df['makeup'].sum() + salary_df['bartending'].sum() + salary_df['cheer'].sum()
        elif (best and bbar and bcheer and not bmakeup):
            salary = salary_df['esthetician'].sum() + salary_df['bartending'].sum() + salary_df['cheer'].sum()
        elif (best and bbar and bmakeup and not bcheer):
            salary = salary_df['esthetician'].sum() + salary_df['bartending'].sum() + salary_df['makeup'].sum()

        elif (bmakeup and best and not bcheer and not bbar):
            salary = salary_df['makeup'].sum() + salary_df['esthetician'].sum() 
        elif (bmakeup and bbar and not bcheer and not best):
            salary = salary_df['makeup'].sum() + salary_df['bartending'].sum()
        elif (not best and not bbar and bcheer and bmakeup):
            salary = salary_df['makeup'].sum() + salary_df['cheer'].sum()

        elif (best and not bbar and bcheer and not bmakeup):
            salary = salary_df['esthetician'].sum() + salary_df['cheer'].sum()
        elif (best and bbar and not bcheer and not bmakeup):
            salary = salary_df['esthetician'].sum() + salary_df['bartending'].sum()

        elif (not best and bbar and bcheer and not bmakeup):
            salary = salary_df['bartending'].sum() + salary_df['cheer'].sum()

        elif (bmakeup and not best and not bbar and not bcheer):
            salary = salary_df['makeup'].sum()
        elif (best and not bbar and not bmakeup and not bcheer):
            salary = salary_df['esthetician'].sum()
        elif (bbar and not best and not bmakeup and not bcheer):
            salary = salary_df['bartending'].sum()
        elif (bcheer and not bbar and not best and not bmakeup):
            salary = salary_df['cheer'].sum()

    return str(salary)

# Function to save Excel file
def save_excel(data):
    # Create a DataFrame from the HTML table data
    try:
        df = pd.DataFrame(data[1:], columns = [data[0][0], data[0][1], data[0][2], data[0][3], data[0][4], data[0][5], data[0][6]])
        df.reset_index(inplace=True, drop=True)
    except:
        df = pd.DataFrame(columns = ['entries', 'makeup', 'bartending', 'esthetician', 'cheer', 'date', 'year'])

    try:
        df.drop(columns={'Unnamed: 0'}, inplace=True)
    except:
        pass

    # File path for the Excel file in the same folder as the script
    excel_file_path = os.path.join(os.path.dirname(__file__), 'salary_data.xlsx')

    # Save the DataFrame to an Excel file
    df.to_excel(excel_file_path, index=False)

if __name__ == "__main__":
    app.run()